import os
import pandas as pd
import logging
from fastapi import HTTPException
import dictionaries as dict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
from openpyxl.utils import get_column_letter


# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


#global variables
idx_start = 0
idx_end = 0
chargeback_data = pd.DataFrame() 
chargeback_data_calculated_columns = pd.DataFrame() 



 
# Excel file processing functions

#Function to process raw data
def process_raw_data(df,idx_start, idx_end):
                
        if df.empty:
            logger.error("Excel sheet is empty")
            raise HTTPException(status_code=400, detail="Excel sheet is empty")
        
        
        # Apply the row processing function to the dataframe
        df = df.apply(lambda row: function_row(row, df, idx_start, idx_end), axis=1)
    
        # Replace infinite values with NaN and fill NaN with 0
        df.replace([float('inf'), -float('inf')], float('nan'), inplace=True)
        df.fillna(0, inplace=True)

        df.columns = df.columns.str.replace('.1', '')

        # Rename the columns to remove the 'Ltd.' abbreviation for consistency
        df.columns = df.columns.str.replace('Ltd.', 'Ltd')

        return df
# Function to process each row
def function_row(row, df, index_start, index_end):
    total_price = row['2024 USD ']  # Direct reference to the column
    sum_value = 0
    diff = index_end - index_start + 1

    for i in range(index_start, index_end + 1):
        col_name = df.columns[i + diff]
        row[col_name] = total_price * row[df.columns[i]]
        sum_value += row[col_name]
    
    row['Control'] = "True" if sum_value == total_price else "False"
    return row





# Sheet styling functions

# Function to style the width of the sheet
def style_sheet_width(writer, sheet_name):
    work_sheet = writer.sheets[sheet_name]
    for col in work_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
            adjusted_width = (max_length + 2)
            work_sheet.column_dimensions[column].width = adjusted_width
    
    return work_sheet

# Function to style the numbers in the sheet
def style_number(writer, sheet_name):
    try:
        # Access the worksheet
        worksheet = writer.sheets[sheet_name]

        # Loop over all rows and columns in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                # Check if the cell contains a number (either int or float)
                if isinstance(cell.value, (int, float)):
                    # Format the number as millions
                    cell.number_format = '#,##0.00,, "M"'  # This formats the number to millions with two decimal points

        logger.info(f"Number formatting applied to sheet: {sheet_name}")

    except Exception as e:
        logger.error(f"Error styling numbers in sheet: {sheet_name}. Error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error styling numbers in sheet: {sheet_name}")





# Functions to get processed file

# Function to save the uploaded file
def save_uploaded_file(upload_file, directory="uploads"):
    try:
        file_location = os.path.join(directory, upload_file.filename)
        with open(file_location, "wb+") as file_object:
            content = upload_file.file.read()
            file_object.write(content)
        logger.info(f"File saved at {file_location}")
        return upload_file.filename
    except Exception as e:
        logger.error(f"Error uploading file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error uploading file: {str(e)}")

# Function to process the Excel file
def process_excel_file(file_path):
    if not os.path.exists(file_path):
        logger.error("File not found: %s", file_path)
        raise HTTPException(status_code=404, detail="File not found")
    try:
        # Read the Excel file
        df = pd.read_excel(file_path, sheet_name='Chargeback Raw Data', skiprows=4)

        # Define the start and end columns
        start_column = 'Beko Plc'
        end_column = 'Arçelik A.Ş.'

        global idx_start
        global idx_end

        # Get the index positions of these columns
        idx_start = df.columns.get_loc(start_column)
        idx_end = df.columns.get_loc(end_column)

        # Process the raw data
        df = process_raw_data(df, idx_start, idx_end)

        global chargeback_data
        chargeback_data = df

        global chargeback_data_calculated_columns
        chargeback_data_calculated_columns = df.iloc[:, idx_end + 1:idx_end + 1 + (idx_end - idx_start + 1)]
        

        # Save the processed DataFrame and summary to an Excel file
        processed_file_path = os.path.join("uploads", f"processed_{os.path.basename(file_path)}")
        writer = pd.ExcelWriter(processed_file_path, engine='openpyxl')
        with writer:
            # Add the main chargeback data sheet
            df.to_excel(writer, index=False, header=True, sheet_name="Chargeback Main")

            # Add the summary sheet
            summary_df = summary_companies()
            summary_df.to_excel(writer, index=False, header=True, sheet_name="Summary")

            # Add the region-based distribution sheet
            region_distribution_df = region_based_distribution(writer)
            region_distribution_df.to_excel(writer, index=False, header=True, sheet_name="Region Based Distribution")
            
            # Add the main categories distribution sheet
            main_categories_df = categorie_based_distributions(writer)
            main_categories_df.to_excel(writer, index=False, header=True, sheet_name="Category Distribution")
            
            # Add the product/service-based distribution sheet
            product_service_distribution_df = product_service_based_distribution(writer)
            product_service_distribution_df.to_excel(writer, index=False, header=True, sheet_name="Product Service Distribution")

            # Add the vendor-based distribution sheet
            vendor_distribution_df = vendor_based_distribution(writer)
            vendor_distribution_df.to_excel(writer, index=False, header=True, sheet_name="Vendor Distribution")

            # Styling of Chargeback Main sheet
            style_sheet_width(writer, "Chargeback Main")
            style_number(writer, "Chargeback Main")



            # Styling of the summary sheet
            style_sheet_width(writer, "Summary")
            work_summary = writer.sheets["Summary"]
            for row in work_summary.iter_rows():
                        for cell in row:
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.000'
            
            
            # Style the header row (A1 to F1)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill_dark = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Dark blue color
            header_fill_brown = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")  # Brown color

            # Apply styles to the cells
            for cell in ['A1', 'B1', 'C1']:
                work_summary[cell].font = header_font
                work_summary[cell].fill = header_fill_dark

            for cell in ['D1', 'E1', 'F1']:
                work_summary[cell].font = header_font
                work_summary[cell].fill = header_fill_brown


            # Styling of Region Based Distribution sheet
            work_region = writer.sheets["Region Based Distribution"]
            # Format the values in the columns in millions for better readability
            style_sheet_width(writer, "Region Based Distribution")
            style_number(writer, "Region Based Distribution")

            
            # Explicitly apply color to the cells A1 dark & B1 brown
            work_region['A1'].font = header_font
            work_region['A1'].fill = header_fill_dark
            work_region['B1'].font = header_font
            work_region['B1'].fill = header_fill_brown

            # Styling of Main Category Distribution sheet
            work_category = writer.sheets["Category Distribution"]
            # Format the values in the columns in millions for better readability
            style_number(writer, "Category Distribution")
            style_sheet_width(writer, "Category Distribution")

            # Explicitly apply color to the cells A1 dark & B1 brown
            work_category['A1'].font = header_font
            work_category['A1'].fill = header_fill_dark
            work_category['B1'].font = header_font
            work_category['B1'].fill = header_fill_brown

            # Styling of Product Service Distribution sheet
            work_product = writer.sheets["Product Service Distribution"]
            # Format the values in the columns in millions for better readability
            style_sheet_width(writer, "Product Service Distribution")
            style_number(writer, "Product Service Distribution")

            # Explicitly apply color to the cells A1 dark & B1 brown
            work_product['A1'].font = header_font
            work_product['A1'].fill = header_fill_dark
            work_product['B1'].font = header_font
            work_product['B1'].fill = header_fill_brown

            # Styling of Vendor Distribution sheet
            work_vendor = writer.sheets["Vendor Distribution"]
            # Format the values in the columns in millions for better readability
            style_number(writer, "Vendor Distribution")
            style_sheet_width(writer, "Vendor Distribution")

        writer.close

        logger.info(f"File processed and saved at {processed_file_path}")
        return {"data": "success", "fileName": os.path.basename(processed_file_path), "result": df.head().to_dict()}
    
    except Exception as e:
        logger.error(f"Error processing file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

# Function to summarize the companies
def summary_companies():
    if chargeback_data_calculated_columns.empty:
        logger.error("Dataframe is empty")
        raise HTTPException(status_code=400, detail="Dataframe is empty")
    try:
        # Create a list to store the summary data before converting to a DataFrame
        summary_rows = []
        
        # Get the list of regions
        regions_dict = dict.get_regions_dict()

        # Getting the calculated columns as a separate dataframe
        calculated_columns = chargeback_data_calculated_columns

        # Get the company-country dictionary
        company_country_dict = dict.get_company_country_dict()

        # Get the sum of each company from the calculated columns
        for region, companies in regions_dict.items():
            for company in companies:
                if company in calculated_columns.columns:
                    company_total = calculated_columns[company].sum()

                    # Log each company's total to verify the sum
                    logger.info(f"Company: {company}, Total: {company_total}")

                    if pd.isna(company_total) or company_total == 0:
                        company_total = 0
                        logger.warning(f"Company: {company} has a zero or NaN total.")

                    # Get the country of the company
                    country = company_country_dict.get(company, "Unknown")

                    # Append each row of data as a dictionary to the list
                    summary_rows.append({
                        'Country': country,
                        'Company': company,
                        'Region': region,
                        '2024 USD (Excl. Markup)': company_total,
                        'Mark up %5': company_total * 0.05,
                        '2024 USD (Inc. Markup)': company_total * 1.05
                    })
        
        # Convert the list of dictionaries to a DataFrame
        summary_df = pd.DataFrame(summary_rows)

        return summary_df
    except Exception as e:
        logger.error(f"Error summarizing companies: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error summarizing companies: {str(e)}")

# Function to get region based on the file processed
def region_based_distribution(writer):
    if chargeback_data.empty:
        logger.error("Dataframe is empty")
        raise HTTPException(status_code=400, detail="Dataframe is empty")

    try:
        # Get the list of regions
        regions_dict = dict.get_regions_dict()
    
        # Getting the calculated columns as a separate dataframe
        calculated_columns = chargeback_data_calculated_columns

        region_data_rows = []

        for region, companies in regions_dict.items():
            region_total = 0
            for company in companies:
                if company in calculated_columns.columns:
                    region_total += calculated_columns[company].sum()
            if region_total > 0:  # Only add regions with a non-zero total
                region_data_rows.append({
                    'Region': region,
                    'Total': region_total
                })

        region_data_df = pd.DataFrame(region_data_rows)

        # Sort the data by Total for better visualization
        region_data_df.sort_values(by='Total', ascending=True, inplace=True)

        # Plot the region-based distribution
        total_value = region_data_df['Total'].sum()

        plt.figure(figsize=(10, 6))
        bars = plt.barh(region_data_df['Region'], region_data_df['Total'], color='blue')
        plt.xlabel('Total ($)')
        plt.title('Region Based Distribution\n(M, $)', pad=20)
        plt.grid(axis='x', linestyle='--', alpha=0.6)

        # Append percentage values at the end and actual values at the center of each bar
        for bar in bars:
            width = bar.get_width()
            percentage = (width / total_value) * 100
            value_in_millions = width / 1_000_000
            
            # Position the value in millions at the center of the bar
            plt.text(width / 2, bar.get_y() + bar.get_height() / 2, 
                     f'{value_in_millions:.2f}M',
                     va='center', ha='center', color='white', fontsize=10)
            
            # Position the percentage at the end of the bar
            plt.text(width, bar.get_y() + bar.get_height() / 2, 
                     f'{percentage:.2f}%',
                     va='center', ha='left', color='black', fontsize=10)

        plt.tight_layout()

        # Save the plot as an image
        img_path = "region_based_distribution.png"
        plt.savefig(img_path)
        plt.close()

        # Insert the region data into a new sheet
        region_data_df.to_excel(writer, index=False, header=True, sheet_name="Region Based Distribution")

        # Load and insert the image into the "Region Based Distribution" sheet
        worksheet = writer.sheets["Region Based Distribution"]
        img = Image(img_path)
        worksheet.add_image(img, 'J5')  # Insert image at cell H2 or any preferred location

        logger.info("Visual inserted into the Excel sheet successfully.")
        return region_data_df

    except Exception as e:
        logger.error(f"Error region based distribution: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error region based distribution: {str(e)}")

    
#Function to visualize Main Categories total and distribution with Pie Chart
def categorie_based_distributions(writer):
    if chargeback_data.empty:
        logger.error("Dataframe is empty")
        raise HTTPException(status_code=400, detail="Dataframe is empty")   
    try:
        # Get the unique main categories from the dataframe
        main_categories = chargeback_data["Main Category"].unique()
        category_data_rows = []
        
        for category in main_categories:
            # Sum the total for each category
            category_total = chargeback_data[chargeback_data["Main Category"] == category]["2024 USD "].sum()
            # Make the values in millions
            category_data_rows.append({
                'Main Category': category,
                '2024 Chargeback Amount': category_total
            })
        
        # Convert the list of dictionaries into a DataFrame
        category_data_df = pd.DataFrame(category_data_rows)

        # Sort the DataFrame by 'Total' in descending order
        category_data_df.sort_values(by='2024 Chargeback Amount', ascending=False, inplace=True)

        # Extract labels and values for the pie chart
        labels = category_data_df['Main Category'].tolist()
        values = category_data_df['2024 Chargeback Amount'].tolist()

        # Convert values to millions
        values_in_millions = [value / 1_000_000 for value in values]

        # Custom label function to display both percentage and actual value in millions
        def func(pct, all_values):
            absolute = int(pct/100.*sum(all_values))
            return f"{absolute/1_000_000:.1f}M\n({pct:.1f}%)"

        # Plot the main categories distribution as a pie chart
        plt.figure(figsize=(8, 6))
        plt.pie(values_in_millions, labels=labels, autopct=lambda pct: func(pct, values), startangle=140)

        # Add a title
        plt.title('Main Category Based Distribution\n(M, $)', pad=20)

        # Save the plot as an image
        img_path = "main_category_distribution.png"
        plt.savefig(img_path)
        plt.close()

        category_data_df.to_excel(writer, index=False, header=True, sheet_name="Category Distribution")

        # Load and insert the image into the "Category Distribution" sheet
        worksheet = writer.sheets["Category Distribution"]
        img = Image(img_path)
        worksheet.add_image(img, 'J5')  # Insert image at cell 

        logger.info("Visual inserted into the Excel sheet successfully.")


        # Group by "Main Category" and "Category 1" to get the totals
        subcategory_totals = chargeback_data.groupby(["Main Category", "Category 1"])["2024 USD "].sum().reset_index()

        # Calculate the total for each "Main Category"
        main_category_totals = chargeback_data.groupby("Main Category")["2024 USD "].sum().reset_index()

        # Write to Excel, starting after the existing data
        sheet_name = "Category Distribution"
        worksheet = writer.sheets[sheet_name]

        # Define row and column start positions
        start_row = len(category_data_df) + 5
        col1 = 1
        col2 = 2

        # Write header
        worksheet.cell(row=start_row, column=col1, value="Category Groups").font = Font(bold=True, color="FF0000")
        worksheet.cell(row=start_row, column=col2, value="Total Chargeback Amount").font = Font(bold=True)
        start_row += 1

        # Set border style for later use
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        grand_total = 0

        # Write each Main Category and its subcategories
        for main_cat in main_category_totals.itertuples():
            # Write Main Category in bold with a total
            worksheet.cell(row=start_row, column=col1, value=main_cat._1).font = Font(bold=True)
            worksheet.cell(row=start_row, column=col2, value=f"{main_cat._2:,.3f}").font = Font(bold=True)
            worksheet.cell(row=start_row, column=col1).border = thin_border
            worksheet.cell(row=start_row, column=col2).border = thin_border
            grand_total += main_cat._2
            start_row += 1

            # Write each subcategory under the main category
            for sub_cat in subcategory_totals[subcategory_totals["Main Category"] == main_cat._1].itertuples():
                worksheet.cell(row=start_row, column=col1, value=sub_cat._2).alignment = Alignment(indent=2)
                worksheet.cell(row=start_row, column=col2, value=f"{sub_cat._3:,.3f}")
                worksheet.cell(row=start_row, column=col1).border = thin_border
                worksheet.cell(row=start_row, column=col2).border = thin_border
                start_row += 1

            start_row += 1  # Add an extra row space between main categories

        # Write the Grand Total at the end
        worksheet.cell(row=start_row, column=col1, value="Grand Total").font = Font(bold=True)
        worksheet.cell(row=start_row, column=col2, value=f"{grand_total:,.3f}").font = Font(bold=True)
        worksheet.cell(row=start_row, column=col1).border = thin_border
        worksheet.cell(row=start_row, column=col2).border = thin_border

        logger.info("Main category distribution and subcategory details written to Excel successfully.")

        return category_data_df

    except Exception as e:
        logger.error(f"Error visualizing main categories: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error visualizing main categories: {str(e)}")

# Function to visualize Product/Service Groups total and distribution 
def product_service_based_distribution(writer):
    if chargeback_data.empty:
        logger.error("Dataframe is empty")
        raise HTTPException(status_code=400, detail="Dataframe is empty")   
    try:
        # Group by "Product/Service Group" to get the totals
        group_totals = chargeback_data.groupby("Product/Service Group")["2024 USD "].sum().reset_index()

        group_totals.to_excel(writer, index=False, header=True, sheet_name="Product Service Distribution")

        worksheet = writer.sheets["Product Service Distribution"]

        # Define row and column start positions
        start_row = 1
        col1 = 1
        col2 = 2

        # Write header
        worksheet.cell(row=start_row, column=col1, value="Product/Service Groups").font = Font(bold=True, color="FF0000")
        worksheet.cell(row=start_row, column=col2, value="Total Chargeback Amount").font = Font(bold=True)
        start_row += 1

        # Set border style for later use
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        grand_total = 0

        # Change the column names to be more readable
        group_totals.columns = ['Product/Service Group', 'Total Chargeback Amount']

        # Write each Product/Service Group and its total
        for group in group_totals.itertuples():
            # Write Product/Service Group in bold with a total
            worksheet.cell(row=start_row, column=col1, value=group._1).font = Font(bold=True)
            worksheet.cell(row=start_row, column=col2, value=f"{group._2:,.3f}").font = Font(bold=True)
            worksheet.cell(row=start_row, column=col1).border = thin_border
            worksheet.cell(row=start_row, column=col2).border = thin_border
            grand_total += group._2
            start_row += 1

        # Write the Grand Total at the end
        worksheet.cell(row=start_row, column=col1, value="Grand Total").font = Font(bold=True)
        worksheet.cell(row=start_row, column=col2, value=f"{grand_total:,.3f}").font = Font(bold=True)
        worksheet.cell(row=start_row, column=col1).border = thin_border
        worksheet.cell(row=start_row, column=col2).border = thin_border

        logger.info("Product/Service distribution details written to Excel successfully.")

        group_totals = group_totals.sort_values(by='Total Chargeback Amount', ascending=False)
        
        labels = group_totals['Product/Service Group']
        values = group_totals['Total Chargeback Amount']

        total_sum = sum(values)
        threshold = 0.1 * total_sum  # 10% threshold
        other_total = 0

        filtered_labels = []
        filtered_values = []

        for label, value in zip(labels, values):
            if value >= threshold:
                filtered_labels.append(label)
                filtered_values.append(value)
            else:
                other_total += value

        if other_total > 0:
            filtered_labels.append('Others')
            filtered_values.append(other_total)

        values_in_millions = [value / 1_000_000 for value in filtered_values]

        # Custom label function to display both percentage and actual value in millions
        def func(pct, all_values):
            absolute = int(pct/100.*sum(all_values))
            return f"{absolute/1_000_000:.1f}M\n({pct:.1f}%)"

        # Plot the product/service groups distribution as a pie chart
        plt.figure(figsize=(8, 6))
        plt.pie(values_in_millions, labels=filtered_labels, autopct=lambda pct: func(pct, filtered_values), startangle=140)

        # Add a title
        plt.title('Product/Service Group Based Distribution\n(M, $)', pad=20)

        # Save the plot as an image
        img_path = "product_service_distribution.png"
        plt.savefig(img_path)
        plt.close()

        # Load and insert the image into the "Product Service Distribution" sheet
        worksheet = writer.sheets["Product Service Distribution"]
        img = Image(img_path)
        worksheet.add_image(img, 'J5')  # Insert image at cell

        logger.info("Visual inserted into the Excel sheet successfully.")

        return group_totals

    except Exception as e:
        logger.error(f"Error visualizing product and service categories: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error visualizing product and service categories: {str(e)}")

# Function to visualize Vendor Groups total and distribution
def vendor_based_distribution(writer):
    if chargeback_data.empty:
        logger.error("Dataframe is empty")
        raise HTTPException(status_code=400, detail="Dataframe is empty")   
    try:
        # Group by "Vendor" to get the totals
        vendor_totals = chargeback_data.groupby("Vendor")["2024 USD "].sum().reset_index()

        # Change the column name to be more readable
        vendor_totals.columns = ['Vendor', 'Total Chargeback Amount']

        # Sort the DataFrame by 'Total Chargeback Amount' in descending order
        vendor_totals.sort_values(by='Total Chargeback Amount', ascending=False, inplace=True)

        vendor_totals.to_excel(writer, index=False, header=True, sheet_name="Vendor Distribution")

        worksheet = writer.sheets["Vendor Distribution"]

        # Define row and column start positions
        start_row = 1
        col1 = 1
        col2 = 2

        # Write header
        worksheet.cell(row=start_row, column=col1, value="Vendor").font = Font(bold=True, color="FF0000")
        worksheet.cell(row=start_row, column=col2, value="Total Chargeback Amount").font = Font(bold=True)
        start_row += 1

        # Set border style for later use
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        grand_total = 0

        # Write each Vendor and its total
        for vendor in vendor_totals.itertuples(index=False, name=None):
            # Access the fields using positional indexing
            worksheet.cell(row=start_row, column=col1, value=vendor[0]).font = Font(bold=True)
            worksheet.cell(row=start_row, column=col2, value=f"{vendor[1]:,.3f}").font = Font(bold=True)
            worksheet.cell(row=start_row, column=col1).border = thin_border
            worksheet.cell(row=start_row, column=col2).border = thin_border
            grand_total += vendor[1]
            start_row += 1

        # Write the Grand Total at the end
        worksheet.cell(row=start_row, column=col1, value="Grand Total").font = Font(bold=True)
        worksheet.cell(row=start_row, column=col2, value=f"{grand_total:,.3f}").font = Font(bold=True)
        worksheet.cell(row=start_row, column=col1).border = thin_border
        worksheet.cell(row=start_row, column=col2).border = thin_border

        # Write the Top 10 Vendors
        top_vendors = vendor_totals.head(10)
        start_row = 1  
        col1 = 4
        col2 = 5

        # Write header for the Top 10 Vendors
        header_fill_yellow = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # Dark yellow color
        worksheet.cell(row=start_row, column=col1, value="Top 10 Vendor").font = Font(bold=True)
        worksheet.cell(row=start_row, column=col2, value="Total Chargeback Amount").font = Font(bold=True)
        worksheet.cell(row=start_row, column=col1).fill = header_fill_yellow
        worksheet.cell(row=start_row, column=col2).fill = header_fill_yellow
        start_row += 1

        # Write each Top 10 Vendor and its total
        for vendor in top_vendors.itertuples(index=False, name=None):
            worksheet.cell(row=start_row, column=col1, value=vendor[0]).font = Font(bold=True)
            worksheet.cell(row=start_row, column=col2, value=f"{vendor[1]:,.3f}").font = Font(bold=True)
            worksheet.cell(row=start_row, column=col1).border = thin_border
            worksheet.cell(row=start_row, column=col2).border = thin_border
            start_row += 1

        # Create the bar chart
        plt.figure(figsize=(10, 6))
        bars = plt.barh(top_vendors['Vendor'], top_vendors['Total Chargeback Amount'] / 1_000_000, color='blue')
        plt.xlabel('Chargeback Amount (M, $)')
        plt.title('Top 10 Vendor List\n(M, $)', pad=20)
        plt.gca().invert_yaxis()  # Invert the y-axis to have the largest value at the top
        plt.tight_layout()

        # Append percentage values at the end and actual values at the center of each bar
        total_value = top_vendors['Total Chargeback Amount'].sum()

        for bar in bars:
            width = bar.get_width()
            percentage = (width * 1_000_000 / total_value) * 100
            value_in_millions = width

            # Position the value in millions at the center of the bar
            plt.text(width / 2, bar.get_y() + bar.get_height() / 2, 
                     f'{value_in_millions:.2f}M',
                     va='center', ha='center', color='white', fontsize=10)
            
            # Position the percentage at the end of the bar
            plt.text(width, bar.get_y() + bar.get_height() / 2, 
                     f'{percentage:.2f}%',
                     va='center', ha='left', color='black', fontsize=10)

        # Save the plot as an image
        img_path = "top_10_vendor_distribution.png"
        plt.savefig(img_path)
        plt.close()

        # Load and insert the image into the "Vendor Distribution" sheet
        img = Image(img_path)
        worksheet.add_image(img, 'J5')

        logger.info("Vendor distribution details written to Excel successfully.")

        return vendor_totals

    except Exception as e:
        logger.error(f"Error visualizing vendor categories: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error visualizing vendor categories: {str(e)}")



# Function to create a subsidiary Excel file

# Function to truncate sheet names to 31 characters without changing any characters
def truncate_sheet_name(name, max_length=31):
    return name[:max_length]

# Function the get dataframe of each subsidiary
def get_subsidaries_df(subsidiary):
    subsidiary_df = pd.DataFrame()

    if chargeback_data.empty:
        logger.error("Dataframe is empty")
        raise HTTPException(status_code=400, detail="Dataframe is empty")
    try:
        # Getting the calculated columns as a separate dataframe
        subsidiary_df = chargeback_data.copy()
        # Identify and rename the first occurrence of the column with the same name as 'subsidiary'
        columns = list(subsidiary_df.columns)
        new_columns = []
        subsidiary_renamed = False

        for col in columns:
            if col == subsidiary and not subsidiary_renamed:
                new_columns.append(f'{subsidiary}_renamed')
                subsidiary_renamed = True  # Rename only the first occurrence
            else:
                new_columns.append(col)

        # Apply the new column names to the DataFrame
        subsidiary_df.columns = new_columns

        # Now select the columns, including both the renamed and the original
        subsidiary_df = subsidiary_df[['Responsible', 'Budget Name', 'Main Category', 'Category 1', 'Category 2', 'Product/Service Group', 'Product', 'Distribution Method', subsidiary]]

        # Change the column name subsidiary to the 2024 IT Chargeback
        subsidiary_df.rename(columns={subsidiary: '2024 IT Chargeback'}, inplace=True)
        # Change the column name Distribution Method to the Chargeback Calculation Method
        subsidiary_df.rename(columns={'Distribution Method': 'Chargeback Calculation Method'}, inplace=True)


        # Drop rows with 0 values in the '2024 IT Chargeback' column
        subsidiary_df = subsidiary_df[subsidiary_df['2024 IT Chargeback'] != 0]

        # Add a column to the beginning of the DataFrame named Company and fill it with the subsidiary name
        subsidiary_df.insert(0, 'Company', subsidiary)

        # Reset the index
        subsidiary_df.reset_index(drop=True, inplace=True)

        return subsidiary_df
    except Exception as e:
        logger.error(f"Error getting subsidiary dataframe: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error getting subsidiary dataframe: {str(e)}")

# Function to create subsidiaries Excel file
def create_subsidaries_excel(file_path):
    if not os.path.exists(file_path):
        logger.error("File not found: %s", file_path)
        raise HTTPException(status_code=404, detail="File not found")

    try:
        calculated_columns = chargeback_data_calculated_columns

        subsidiaries = calculated_columns.columns.tolist()  
        # Create a new Excel file
        subsidiaries_file_path = os.path.join("uploads", f"subsidiaries_{os.path.basename(file_path)}")

        with pd.ExcelWriter(subsidiaries_file_path, engine='openpyxl') as writer:
            for subsidiary in subsidiaries:
                truncated_name = truncate_sheet_name(subsidiary) 
                subsidiary_df = get_subsidaries_df(subsidiary)
                if not subsidiary_df.empty:  # Check if the DataFrame is not empty
                    subsidiary_df.to_excel(writer, index=False, header=True, sheet_name=truncated_name)
                    total_value = subsidiary_df['2024 IT Chargeback'].sum()
                    last_row_data = {
                        '2024 IT Chargeback': total_value,
                        '2024 IT Chargeback Markup %5': total_value * 0.05,
                        '2024 IT Chargeback Inc. Markup': total_value * 1.05
                    }
                    last_row = pd.DataFrame([last_row_data])
                    #write it Columns H,I,J
                    last_row.to_excel(writer, index=False, header=True, sheet_name=truncated_name, startrow=len(subsidiary_df) + 1, startcol=7)
                    # Style the sheet
                    style_sheet_width(writer, truncated_name)
                    work = writer.sheets[truncated_name]
                    for row in work.iter_rows():
                        for cell in row:
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.000'
                    

                    worksheet = writer.sheets[truncated_name]
                    header_fill_yellow = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                    row = len(subsidiary_df) + 2
                    for col in range(8, 11):
                        worksheet.cell(row=row, column=col).fill = header_fill_yellow
                    header_fill_blue = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") 
                    for col in range(1, 11):
                        worksheet.cell(row=1, column=col).fill = header_fill_blue
                    


            writer.close

        if os.path.exists(subsidiaries_file_path):
            logger.info(f"Subsidiaries Excel file created at {subsidiaries_file_path}")
        else:
            logger.error(f"Failed to create the file at {subsidiaries_file_path}")

        return subsidiaries_file_path

    except Exception as e:
        logger.error(f"Error creating subsidiaries Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error creating subsidiaries Excel file: {str(e)}")




#Functions to retrieve available regions and vendors
def get_available_regions(file_path):
    if not os.path.exists(file_path):
        logger.error("File not found: %s", file_path)
        raise HTTPException(status_code=404, detail="File not found")

    try:
        df = pd.read_excel(file_path, sheet_name='Chargeback Raw Data', skiprows=4)
        
        regions_dict = dict.get_regions_dict()
        
        available_regions = []
        for region, countries in regions_dict.items():
            total_price = sum(df[country].sum() for country in countries if country in df.columns)
            if total_price > 0:
                available_regions.append(region)
        
        return available_regions
    except Exception as e:
        logger.error(f"Error retrieving available regions: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error retrieving available regions: {str(e)}")
def get_available_vendors(file_path):
    if not os.path.exists(file_path):
        logger.error("File not found: %s", file_path)
        raise HTTPException(status_code=404, detail="File not found")

    try:
        df = pd.read_excel(file_path, sheet_name='Chargeback Raw Data', skiprows=4)

        vendors_dict = df.groupby("Vendor")["2024 TRY"].sum().to_dict()
        return list(vendors_dict.keys())
    except Exception as e:
        logger.error(f"Error retrieving available vendors: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error retrieving available vendors: {str(e)}")


    

